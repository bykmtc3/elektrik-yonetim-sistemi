from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime
import os
from config import Config

class ReportGenerator:
    def __init__(self):
        self.report_dir = Config.REPORT_DIR
        os.makedirs(self.report_dir, exist_ok=True)
    
    def generate_excel_report(self, project_name, materials_data):
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Malzeme Listesi"
        
        # Headers
        ws['A1'] = "Proje Adı"
        ws['B1'] = project_name
        ws['A2'] = "Tarih"
        ws['B2'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Material headers
        ws['A4'] = "Malzeme"
        ws['B4'] = "Miktar"
        ws['C4'] = "Birim"
        ws['D4'] = "Birim Fiyat"
        ws['E4'] = "Toplam"
        
        # Data
        row = 5
        total = 0
        for item in materials_data:
            ws[f'A{row}'] = item.get('material', '')
            ws[f'B{row}'] = item.get('quantity', 0)
            ws[f'C{row}'] = item.get('unit', '')
            ws[f'D{row}'] = item.get('unit_price', 0)
            ws[f'E{row}'] = item.get('cost', 0)
            total += item.get('cost', 0)
            row += 1
        
        # Total
        ws[f'D{row}'] = "TOPLAM:"
        ws[f'E{row}'] = total
        
        # Save
        filename = f"Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        wb.save(filepath)
        return filepath
    
    def generate_pdf_report(self, project_name, materials_data):
        """Generate PDF report"""
        filename = f"Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.report_dir, filename)
        
        c = canvas.Canvas(filepath, pagesize=letter)
        
        # Title
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, 750, f"Elektrik Malzeme Raporu - {project_name}")
        
        # Date
        c.setFont("Helvetica", 10)
        c.drawString(50, 730, f"Tarih: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Table
        y = 700
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "Malzeme")
        c.drawString(200, y, "Miktar")
        c.drawString(300, y, "Birim")
        c.drawString(400, y, "Fiyat")
        
        y -= 20
        c.setFont("Helvetica", 9)
        for item in materials_data:
            c.drawString(50, y, item.get('material', ''))
            c.drawString(200, y, str(item.get('quantity', '')))
            c.drawString(300, y, item.get('unit', ''))
            c.drawString(400, y, str(item.get('cost', '')))
            y -= 15
        
        c.save()
        return filepath